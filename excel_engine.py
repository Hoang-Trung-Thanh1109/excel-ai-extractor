"""Excel workbook creation and formatting."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
from datetime import datetime
import os
import re
import subprocess
import sys


@dataclass
class WorkbookBuildResult:
    path: str


class ExcelEngine:
    def __init__(self, logger=None) -> None:
        self.logger = logger

    def save_workbook(self, workbook_spec: dict[str, Any], output_path: str | None = None) -> WorkbookBuildResult:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise RuntimeError(
                "Thieu goi 'openpyxl'. Hay cai dat trong requirements.txt truoc khi chay."
            ) from exc

        workbook_title = str(workbook_spec.get("workbook_title") or "Excel AI Extractor")
        sheets = workbook_spec.get("sheets") or []

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        zebra_fill = PatternFill("solid", fgColor="F7FBFF")
        border_side = Side(style="thin", color="D9E2F3")
        cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        body_alignment = Alignment(vertical="top", wrap_text=True)

        for sheet_index, sheet in enumerate(sheets, start=1):
            if not isinstance(sheet, dict):
                continue

            sheet_name = self._sanitize_sheet_name(str(sheet.get("name") or f"Sheet{sheet_index}"))
            headers = [str(item) for item in (sheet.get("headers") or []) if str(item).strip()]
            rows = sheet.get("rows") or []
            if not headers and isinstance(rows, list) and rows and isinstance(rows[0], dict):
                headers = list(rows[0].keys())
            if not headers:
                headers = ["Column 1"]

            worksheet = workbook.create_sheet(title=sheet_name)
            worksheet.append(headers)
            for row in rows:
                if isinstance(row, dict):
                    worksheet.append([row.get(header, "") for header in headers])
                elif isinstance(row, list):
                    worksheet.append([row[i] if i < len(row) else "" for i in range(len(headers))])
                else:
                    worksheet.append([row])

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = cell_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for row_index, row_cells in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                for cell in row_cells:
                    cell.border = cell_border
                    cell.alignment = body_alignment
                    if row_index % 2 == 0:
                        cell.fill = zebra_fill

            self._auto_width(worksheet, get_column_letter)

        if not workbook.sheetnames:
            worksheet = workbook.create_sheet(title="Sheet1")
            worksheet["A1"] = "Khong co du lieu"

        workbook.properties.title = workbook_title
        if output_path:
            target = Path(output_path)
        else:
            outputs_dir = Path.cwd() / "outputs"
            outputs_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            target = outputs_dir / f"excel_ai_extractor_{timestamp}.xlsx"

        target.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(target))
        if self.logger:
            self.logger(f"Da luu Excel: {target}")
        return WorkbookBuildResult(path=str(target))

    def open_workbook(self, path: str) -> None:
        if not path:
            return

        if sys.platform == "darwin":
            subprocess.Popen(["open", path])
        elif os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
        else:
            subprocess.Popen(["xdg-open", path])

    def _sanitize_sheet_name(self, name: str) -> str:
        cleaned = re.sub(r"[\[\]\*\?/\\:]", "_", name).strip()
        if not cleaned:
            cleaned = "Sheet"
        return cleaned[:31]

    def _auto_width(self, worksheet, get_column_letter) -> None:
        for column_index, column_cells in enumerate(worksheet.columns, start=1):
            max_length = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            column_letter = get_column_letter(column_index)
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 4, 12), 45)
